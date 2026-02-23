"""
create_workbook.py
Generates (or regenerates) the Golf Round Recap Excel base template.
Saves to repo only -- does NOT overwrite the live Google Drive file.
To initialise a fresh Drive file manually copy the output:
  copy "Golf Round Recap.xlsx" "G:\\My Drive\\Project_Outputs\\Golf Round Recap\\Golf Round Recap.xlsx"
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import date
import pathlib

TEMPLATE_PATH = pathlib.Path(__file__).parent / "Golf Round Recap.xlsx"

# ── Styles ────────────────────────────────────────────────────────────────────
HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
HDR_FONT  = Font(color="FFFFFF", bold=True, name="Calibri")
BODY_FONT = Font(name="Calibri", size=11)
ALT_FILL  = PatternFill("solid", fgColor="D9E8F5")

def style_header(cell):
    cell.font = HDR_FONT
    cell.fill = HDR_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def write_row(ws, row_num, values, fill=None):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.font = BODY_FONT
        if fill:
            cell.fill = fill

def dv(formula, sqref):
    v = DataValidation(type="list", formula1=formula, allow_blank=True, showErrorMessage=False)
    v.sqref = sqref
    return v


# ============================================================
# TAB 1 -- ROUNDS  (29 columns)
# ============================================================
#  1  Date               11  Penalties          21  Short Irons (8-P)
#  2  Course             12  FIR                22  Wedges (GW/SW/LW)
#  3  Note Type          13  GIR                23  Putter
#  4  Hole               14  Tee Club           24  Playing Handicap
#  5  Par                15  Pick Up            25  Tee Colour
#  6  Distance (m)       16  Sentiment          26  Course Rating   } Overall
#  7  Stroke Index       17  Driver             27  Slope           } rows
#  8  Score              18  Woods              28  WHS Index       } only
#  9  Strokes            19  Hybrids
# 10  Putts              20  Long Irons (5-7)   29  Notes
# ============================================================
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Rounds"

ROUND_HEADERS = [
    "Date", "Course", "Note Type", "Hole", "Par", "Distance (m)", "Stroke Index",
    "Score", "Strokes", "Putts", "Penalties",
    "FIR", "GIR",
    "Tee Club", "Pick Up", "Sentiment",
    "Driver", "Woods", "Hybrids", "Long Irons\n(5-7)", "Short Irons\n(8-P)", "Wedges\n(GW/SW/LW)", "Putter",
    "Playing\nHandicap", "Tee Colour",
    "Course\nRating", "Slope", "WHS\nIndex",
    "Notes",
]
ROUND_COL_WIDTHS = [
    12, 20, 12, 7, 6, 14, 13,
    16, 10, 8, 11,
    7, 7,
    14, 10, 12,
    11, 10, 11, 13, 13, 13, 10,
    11, 12,
    11, 9, 11,
    60,
]

ws.row_dimensions[1].height = 36
for col, h in enumerate(ROUND_HEADERS, 1):
    style_header(ws.cell(row=1, column=col, value=h))
    ws.column_dimensions[get_column_letter(col)].width = ROUND_COL_WIDTHS[col - 1]

ws.freeze_panes = "A2"

STRIKE = '"Great,Good,Average,Poor"'
ws.add_data_validation(dv('"Overall,Hole"',                                           "C2:C9999"))
ws.add_data_validation(dv('"Eagle,Birdie,Par,Bogey,Double Bogey,Triple Bogey,Other"', "H2:H9999"))
ws.add_data_validation(dv('"Y,N"',                                                    "L2:L9999"))  # FIR
ws.add_data_validation(dv('"Y,N"',                                                    "M2:M9999"))  # GIR
ws.add_data_validation(dv('"Y,N"',                                                    "O2:O9999"))  # Pick Up
ws.add_data_validation(dv('"Positive,Neutral,Negative"',                              "P2:P9999"))  # Sentiment
ws.add_data_validation(dv(STRIKE,                                                     "Q2:Q9999"))  # Driver
ws.add_data_validation(dv(STRIKE,                                                     "R2:R9999"))  # Woods
ws.add_data_validation(dv(STRIKE,                                                     "S2:S9999"))  # Hybrids
ws.add_data_validation(dv(STRIKE,                                                     "T2:T9999"))  # Long Irons
ws.add_data_validation(dv(STRIKE,                                                     "U2:U9999"))  # Short Irons
ws.add_data_validation(dv(STRIKE,                                                     "V2:V9999"))  # Wedges
ws.add_data_validation(dv(STRIKE,                                                     "W2:W9999"))  # Putter
ws.add_data_validation(dv('"White,Yellow,Red,Blue,Black"',                            "Y2:Y9999"))  # Tee Colour

# ── Sample Overall row ────────────────────────────────────────────────────────
write_row(ws, 2, [
    date(2026, 2, 22), "Pupuke", "Overall", 0, 70, 5188, None,
    101, 101, None, None,
    None, None,                              # FIR, GIR -- blank on Overall
    None, None, "Negative",                  # Tee Club, Pick Up, Sentiment
    "Great", "Poor", "Average", "Poor", "Average", "Good", "Average",
    None, "White",
    68.2, 119, 24.4,
    "Tee time 7:22am. Sample overall notes.",
], fill=ALT_FILL)

# ── Sample Hole row ───────────────────────────────────────────────────────────
write_row(ws, 3, [
    date(2026, 2, 22), "Pupuke", "Hole", 4, 4, 304, 13,
    "Par", 4, 2, None,
    "Y", "Y",                                # FIR, GIR
    "Driver", "N", "Positive",
    None, None, None, None, None, None, None,
    None, None,
    None, None, None,                        # Course Rating, Slope, WHS -- blank on Hole rows
    "Good drive, good PW onto green, two putt.",
])


# ============================================================
# TAB 2 -- COURSES  (9 columns)
# ============================================================
#  1  Course       4  Par           7  Course Rating
#  2  Hole         5  Distance (m)  8  Slope
#  3  Tee Colour   6  Stroke Index  9  Notes
# ============================================================
ws_c = wb.create_sheet("Courses")

COURSE_HEADERS = ["Course", "Hole", "Tee Colour", "Par", "Distance (m)",
                  "Stroke Index", "Course Rating", "Slope", "Notes"]
COURSE_COL_WIDTHS = [22, 7, 12, 6, 14, 14, 13, 9, 40]

ws_c.row_dimensions[1].height = 30
for col, h in enumerate(COURSE_HEADERS, 1):
    style_header(ws_c.cell(row=1, column=col, value=h))
    ws_c.column_dimensions[get_column_letter(col)].width = COURSE_COL_WIDTHS[col - 1]

ws_c.freeze_panes = "A2"

# Pupuke Golf Club -- White tees (verified)
PUPUKE = [
    # hole, tee,     par, dist_m, stroke_index
    ( 1, "White",  4,  300,  7),
    ( 2, "White",  3,  139, 15),
    ( 3, "White",  4,  335,  3),
    ( 4, "White",  4,  304, 13),
    ( 5, "White",  5,  431,  9),
    ( 6, "White",  3,  165, 11),
    ( 7, "White",  4,  363,  1),
    ( 8, "White",  4,  333,  5),
    ( 9, "White",  3,  147, 17),
    (10, "White",  5,  422, 14),
    (11, "White",  5,  398, 16),
    (12, "White",  4,  362,  2),
    (13, "White",  3,  167,  8),
    (14, "White",  4,  285, 10),
    (15, "White",  4,  299,  6),
    (16, "White",  4,  238, 18),
    (17, "White",  3,  143, 12),
    (18, "White",  4,  357,  4),
]

for row_i, (hole, tee, par, dist, si) in enumerate(PUPUKE, 2):
    fill = ALT_FILL if row_i % 2 == 1 else None
    write_row(ws_c, row_i, ["Pupuke", hole, tee, par, dist, si, None, None, ""], fill=fill)

total_par  = sum(h[2] for h in PUPUKE)
total_dist = sum(h[3] for h in PUPUKE)
bold = Font(bold=True, name="Calibri")
for col, val in enumerate(["Pupuke", "TOTAL", "White", total_par, total_dist, None, 68.2, 119, ""], 1):
    cell = ws_c.cell(row=20, column=col, value=val)
    cell.font = bold


# ============================================================
# TAB 3 -- GUIDE
# ============================================================
ws_g = wb.create_sheet("Guide")
ws_g.sheet_view.showGridLines = False

SEC_FILL  = PatternFill("solid", fgColor="1F4E79")
SUB_FILL  = PatternFill("solid", fgColor="2E75B6")
SUB_FONT  = Font(color="FFFFFF", bold=True, name="Calibri", size=10)
GUIDE_BODY = Font(name="Calibri", size=11)

ws_g.column_dimensions["A"].width = 32
ws_g.column_dimensions["B"].width = 20
ws_g.column_dimensions["C"].width = 55

def guide_section(row, title):
    ws_g.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    cell = ws_g.cell(row=row, column=1, value=title)
    cell.font = Font(color="FFFFFF", bold=True, name="Calibri", size=12)
    cell.fill = SEC_FILL
    ws_g.row_dimensions[row].height = 20

def guide_subhdr(row, cols):
    for col, val in enumerate(cols, 1):
        cell = ws_g.cell(row=row, column=col, value=val)
        cell.font = SUB_FONT
        cell.fill = SUB_FILL

def guide_row(row, vals):
    for col, val in enumerate(vals, 1):
        cell = ws_g.cell(row=row, column=col, value=val)
        cell.font = GUIDE_BODY
        if row % 2 == 0:
            cell.fill = ALT_FILL

r = 1
guide_section(r, "SCORE LABELS -- Score column on Hole rows"); r += 1
guide_subhdr(r, ["Result", "Strokes vs Par", "Example (par 4)"]); r += 1
for vals in [
    ("Eagle",        "-2 or better",                "2 strokes on a par 4"),
    ("Birdie",       "-1",                           "3 strokes on a par 4"),
    ("Par",          "0",                            "4 strokes on a par 4"),
    ("Bogey",        "+1",                           "5 strokes on a par 4"),
    ("Double Bogey", "+2",                           "6 strokes on a par 4"),
    ("Triple Bogey", "+3",                           "7 strokes on a par 4"),
    ("Other",        "+4 or worse, or pick up",      "8+ strokes or picked up"),
]:
    guide_row(r, vals); r += 1

r += 1
guide_section(r, "FIR -- Fairway in Regulation (par 4s and 5s only, leave blank on par 3s and Overall)"); r += 1
guide_subhdr(r, ["Value", "Meaning", ""]); r += 1
guide_row(r, ["Y", "Tee shot landed on the fairway", ""]); r += 1
guide_row(r, ["N", "Tee shot missed fairway (rough, bunker, OB, etc.)", ""]); r += 1
guide_row(r, ["(blank)", "Par 3 tee shots and Overall rows", ""]); r += 1

r += 1
guide_section(r, "GIR -- Green in Regulation (all hole rows, leave blank on Overall)"); r += 1
guide_subhdr(r, ["Value", "Meaning", "Regulation shots to green"]); r += 1
guide_row(r, ["Y", "Ball on putting surface within regulation", "Par 3 = tee shot | Par 4 = within 2 | Par 5 = within 3"]); r += 1
guide_row(r, ["N", "Missed the green in regulation", "Chipping / pitching in = N"]); r += 1
guide_row(r, ["(blank)", "Pick up holes and Overall rows", ""]); r += 1

r += 1
guide_section(r, "SENTIMENT -- how you felt the hole or round went"); r += 1
guide_subhdr(r, ["Value", "Your words might include...", ""]); r += 1
for vals in [
    ("Positive",  "great, best hole, happy, exactly the plan, nice, solid, holed it, love it", ""),
    ("Neutral",   "ok, sensible, got away with it, fine, recovered, not bad, average",         ""),
    ("Negative",  "disappointed, disaster, terrible, nightmare, duffed, dire, struggled, hack, awful, poor", ""),
]:
    guide_row(r, vals); r += 1

r += 1
guide_section(r, "BALL STRIKING RATINGS -- Overall rows, one rating per club category"); r += 1
guide_subhdr(r, ["Rating", "Your words might include...", "Numeric (PBI)"]); r += 1
for vals in [
    ("Great",   "monster, perfect, very good, flushed it, exactly where I wanted", "4"),
    ("Good",    "good, solid, nice, decent, hit it well",                          "3"),
    ("Average", "ok, bit fadey, slight fade/slice, average, could be better",      "2"),
    ("Poor",    "duffed, hacked, bladed, hooked, sliced, below average, dire",     "1"),
]:
    guide_row(r, vals); r += 1

r += 1
guide_section(r, "SENTIMENT NUMERIC SCALE (for Power BI)"); r += 1
guide_subhdr(r, ["Text", "Numeric", ""]); r += 1
for vals in [("Positive", 5, ""), ("Neutral", 3, ""), ("Negative", 1, "")]:
    guide_row(r, vals); r += 1

r += 1
guide_section(r, "OVERALL ROW TIPS -- one per round, Hole = 0"); r += 1
tips = [
    "Notes: tee time, weather, wind, course conditions, overall impressions.",
    "Score / Strokes = gross total for the round.",
    "FIR / GIR = leave blank on Overall rows (fill totals in notes if wanted).",
    "Course Rating, Slope, WHS Index = persist here so changes over time are captured.",
    "Ball striking = how each club category felt across the whole round.",
    "Putts / Penalties = total for the round.",
]
for tip in tips:
    ws_g.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    cell = ws_g.cell(row=r, column=1, value=tip)
    cell.font = GUIDE_BODY
    r += 1


# ============================================================
# TAB 4 -- HOW TO  (operational workflow -- mirrors README)
# ============================================================
ws_h = wb.create_sheet("How To")
ws_h.sheet_view.showGridLines = False
ws_h.column_dimensions["A"].width = 6   # step number / bullet
ws_h.column_dimensions["B"].width = 28  # heading / label
ws_h.column_dimensions["C"].width = 70  # detail

def ht_section(row, title):
    ws_h.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    cell = ws_h.cell(row=row, column=1, value=title)
    cell.font = Font(color="FFFFFF", bold=True, name="Calibri", size=12)
    cell.fill = SEC_FILL
    ws_h.row_dimensions[row].height = 20

def ht_row(row, step, label, detail=""):
    ws_h.cell(row=row, column=1, value=step).font  = Font(bold=True, name="Calibri", size=11)
    ws_h.cell(row=row, column=2, value=label).font = Font(bold=True, name="Calibri", size=11)
    cell = ws_h.cell(row=row, column=3, value=detail)
    cell.font = GUIDE_BODY
    cell.alignment = Alignment(wrap_text=True)
    if row % 2 == 0:
        for c in range(1, 4):
            ws_h.cell(row=row, column=c).fill = ALT_FILL
    ws_h.row_dimensions[row].height = 28

r = 1

ht_section(r, "FILE LOCATIONS"); r += 1
ht_row(r, "", "Live data file (edit here)",
       r"G:\My Drive\Project_Outputs\Golf Round Recap\Golf Round Recap.xlsx"); r += 1
ht_row(r, "", "Base template (source control)",
       r"C:\Users\craig.f\Home_Projects\Golf Round Recap\Golf Round Recap.xlsx"); r += 1
ht_row(r, "", "GitHub repo",
       "https://github.com/C-Fowler-Tech/golf-round-recap"); r += 1

r += 1
ht_section(r, "ENTERING A NEW ROUND"); r += 1
for step, label, detail in [
    ("1", "Open live file",        r"G:\My Drive\Project_Outputs\Golf Round Recap\Golf Round Recap.xlsx"),
    ("2", "Add Overall row",       "Note Type = Overall, Hole = 0. Fill date, course, tee colour, gross score, "
                                   "playing handicap, course rating, slope, WHS index, ball striking ratings, notes "
                                   "(tee time, weather, conditions, overall impressions)."),
    ("3", "Add Hole rows",         "One row per hole. Note Type = Hole. Fill par, distance, stroke index from "
                                   "the Courses tab. Fill FIR (par 4/5 only), GIR, score, strokes, putts, "
                                   "penalties, tee club, pick up, sentiment, notes."),
    ("4", "Save",                  "OneDrive AutoSave handles sync. No manual save needed."),
    ("5", "Commit to git (optional)", "After a data entry session, commit the Drive file for backup history:\n"
                                   "git add . && git commit -m 'Add round: Pupuke DD-Mon-YYYY' && git push"),
]:
    ht_row(r, step, label, detail); r += 1

r += 1
ht_section(r, "CHANGING THE SCHEMA (adding/removing columns)"); r += 1
for step, label, detail in [
    ("1", "Update create_workbook.py", "Add the new column to ROUND_HEADERS, ROUND_COL_WIDTHS, sample rows, "
                                       "and data validations. Update Guide tab content if needed."),
    ("2", "Run create_workbook.py",    "Saves new template to repo only -- does NOT touch the Drive file.\n"
                                       "cd 'C:\\Users\\craig.f\\Home_Projects\\Golf Round Recap'\n"
                                       "python create_workbook.py"),
    ("3", "Copy template to Drive",    "MUST do this before reloading data, otherwise data lands in wrong columns.\n"
                                       "python -c \"import shutil; shutil.copy2('Golf Round Recap.xlsx', "
                                       r"r'G:\My Drive\Project_Outputs\Golf Round Recap\Golf Round Recap.xlsx')\""),
    ("4", "Reload data",               "Run any data loading scripts (e.g. load_round_YYYYMMDD.py) after "
                                       "the template has been copied. Data will now align to the new schema."),
    ("5", "Update README.md",          "Keep the column guide in README.md in sync with the workbook. "
                                       "Both live in source control -- update together in the same commit."),
    ("6", "Commit everything",         "git add . && git commit -m 'Schema: describe change' && git push"),
]:
    ht_row(r, step, label, detail); r += 1

r += 1
ht_section(r, "ADDING A NEW COURSE"); r += 1
for step, label, detail in [
    ("1", "Open Courses tab",      "Add 18 rows (one per hole) for the new course and tee colour."),
    ("2", "Fill columns",          "Course name must match EXACTLY what you will type in the Rounds tab. "
                                   "Fill Hole, Tee Colour, Par, Distance (m), Stroke Index. "
                                   "On the TOTAL row, fill Course Rating and Slope."),
    ("3", "Update create_workbook.py", "Add the course data to the script so it is included in future template resets."),
]:
    ht_row(r, step, label, detail); r += 1

r += 1
ht_section(r, "POWER BI"); r += 1
for step, label, detail in [
    ("", "Data source",          r"Connect to G:\My Drive\Project_Outputs\Golf Round Recap\Golf Round Recap.xlsx"),
    ("", "Rounds grain",         "Filter Note Type = 'Overall' for round-level measures. "
                                 "Filter Note Type = 'Hole' for hole-level measures."),
    ("", "Sentiment (numeric)",  "Positive = 5, Neutral = 3, Negative = 1  (calculated column in PBI)"),
    ("", "Ball striking (numeric)", "Great = 4, Good = 3, Average = 2, Poor = 1  (calculated column in PBI)"),
    ("", "Handicap differential",  "Formula: (Score - Course Rating) x 113 / Slope"),
]:
    ht_row(r, step, label, detail); r += 1


# ── Save ─────────────────────────────────────────────────────────────────────
wb.save(TEMPLATE_PATH)
print(f"Template saved : {TEMPLATE_PATH}")
print(f"  Rounds: 29 cols | Courses: 9 cols | Guide: updated")
print(f"  Pupuke White: Par {total_par}, {total_dist}m, Rating 68.2, Slope 119")
print(f"  Drive file NOT touched -- copy manually only when a full reset is needed.")
