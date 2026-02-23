"""
One-off data entry script -- Pupuke, 4 Jan 2026.
Delete this script once verified.
"""

import openpyxl
from openpyxl.styles import Font
from datetime import date

FILE = r"G:\My Drive\Project_Outputs\Golf Round Recap\Golf Round Recap.xlsx"
BODY_FONT = Font(name="Calibri", size=11)

wb = openpyxl.load_workbook(FILE)
ws = wb["Rounds"]

# Find next empty row
next_row = ws.max_row + 1
while next_row > 2 and ws.cell(row=next_row - 1, column=1).value is None:
    next_row -= 1

def wr(row_num, values):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.font = BODY_FONT

D = date(2026, 1, 4)

HOLES = [
    # hole, par, dist, SI
    ( 1, 4, 300,  7),
    ( 2, 3, 139, 15),
    ( 3, 4, 335,  3),
    ( 4, 4, 304, 13),
    ( 5, 5, 431,  9),
    ( 6, 3, 165, 11),
    ( 7, 4, 363,  1),
    ( 8, 4, 333,  5),
    ( 9, 3, 147, 17),
    (10, 5, 422, 14),
    (11, 5, 398, 16),
    (12, 4, 362,  2),
    (13, 3, 167,  8),
    (14, 4, 285, 10),
    (15, 4, 299,  6),
    (16, 4, 238, 18),
    (17, 3, 143, 12),
    (18, 4, 357,  4),
]

STROKES = [4, 4, 6, 4, 8, 5, 6, 6, 2, 5, 6, 6, 5, 6, 7, 5, 4, 6]

def score_label(strokes, par, pickup="N"):
    if pickup == "Y":
        return "Pick Up"
    return {"Hole in One": -2, **{-3:"Albatross",-2:"Eagle",-1:"Birdie",0:"Par",
            1:"Bogey",2:"Double Bogey",3:"Triple Bogey"}}.get(strokes - par,
           {-3:"Albatross",-2:"Eagle",-1:"Birdie",0:"Par",
            1:"Bogey",2:"Double Bogey",3:"Triple Bogey"}.get(strokes - par, "Other"))

def score_label(strokes, par, pickup="N"):
    if pickup == "Y":
        return "Pick Up"
    return {-3:"Albatross",-2:"Eagle",-1:"Birdie",0:"Par",
            1:"Bogey",2:"Double Bogey",3:"Triple Bogey"}.get(strokes - par, "Other")

# (tee_club, pickup, sentiment, putts, penalties, fir, gir, notes)
DETAIL = {
    1:  ("Driver", "N", "Neutral",  None, None, None, None, "Par."),
    2:  ("",       "N", "Neutral",  None, None, None, None, "Bogey."),
    3:  ("Driver", "N", "Negative", None, None, None, None, "Double bogey."),
    4:  ("Driver", "N", "Positive", None, None, None, None, "Par."),
    5:  ("Driver", "N", "Negative", None, None, None, None,
         "8 on the par 5 - costly hole on the front nine."),
    6:  ("",       "N", "Negative", None, None, None, None, "Double bogey on par 3."),
    7:  ("Driver", "N", "Negative", None, None, None, None, "Double bogey."),
    8:  ("Driver", "N", "Negative", None, None, None, None, "Double bogey."),
    9:  ("",       "N", "Positive", 1,    None, None, "N",
         "Chip in birdie. Best hole of the round."),
    10: ("Driver", "N", "Neutral",  None, None, None, None, "Par on the par 5."),
    11: ("Driver", "N", "Neutral",  None, None, None, None, "Bogey."),
    12: ("Driver", "N", "Negative", None, None, None, None, "Double bogey."),
    13: ("",       "N", "Negative", None, None, None, None,
         "Double bogey. Duff chip cost a shot."),
    14: ("Driver", "N", "Negative", None, 1,    "N",  None,
         "Penalty off the tee. Double bogey."),
    15: ("Driver", "N", "Negative", None, 1,    "Y",  None,
         "Penalty from 3 hybrid second shot. Triple bogey."),
    16: ("",       "N", "Neutral",  None, None, None, None, "Bogey."),
    17: ("",       "N", "Neutral",  None, None, None, None, "Bogey."),
    18: ("Driver", "N", "Negative", None, None, None, None, "Double bogey."),
}

# ── Overall row ───────────────────────────────────────────────────────────────
wr(next_row, [
    D, "Pupuke", "Overall", 0, 70, 5188, None,
    95, 95, None, 2,
    None, None,
    None, None, "Neutral",
    "Average", None, "Average", "Average", "Average", "Good", "Poor",
    None, "White",
    68.2, 119, 24.4,
    "Course in good condition - firm dry track. "
    "Pretty good round overall but slipped away with a couple of bad holes, penalties and putting. "
    "Good front nine for 45 including chip in birdie on 9, offset by an 8 on the 5th. "
    "Could have been on for breaking 90 - still the next goal. "
    "Driver ok but penalty off tee on 14. Irons ok. Chipping and wedges good. "
    "Putter poor - could not close out second putts, leading to three putts and double bogeys.",
])
next_row += 1

# ── Hole rows ─────────────────────────────────────────────────────────────────
for hole, par, dist, si in HOLES:
    strokes = STROKES[hole - 1]
    tee_club, pickup, sentiment, putts, penalties, fir, gir, notes = DETAIL[hole]
    score = score_label(strokes, par, pickup)
    wr(next_row, [
        D, "Pupuke", "Hole", hole, par, dist, si,
        score, strokes, putts, penalties,
        fir, gir,
        tee_club, pickup, sentiment,
        None, None, None, None, None, None, None,
        None, None,
        None, None, None,
        notes,
    ])
    next_row += 1

wb.save(FILE)
print(f"Round loaded. Rows written up to row {next_row - 1}.")
print("Jan 4 2026 -- Pupuke -- 95 (front 45 / back 50). Delete script once verified.")
