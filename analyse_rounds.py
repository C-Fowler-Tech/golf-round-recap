"""
analyse_rounds.py
Reads Golf Round Recap.xlsx, writes Golf Round Recap - Analysis.xlsx
to the same Google Drive output folder.  Re-run after each new round.

Tabs:
  Round Summary  -- one row per round, calculated fields (vs par, hcp diff)
  Hole Stats     -- one row per hole (course/tee/hole), aggregated across rounds
  Scoring        -- score label distribution with par 3/4/5 breakdown
  Sentiment      -- hole-level sentiment breakdown
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from collections import defaultdict
from datetime import date as dt_date

SRC  = r"G:\My Drive\Project_Outputs\Golf Round Recap\Golf Round Recap.xlsx"
DEST = r"G:\My Drive\Project_Outputs\Golf Round Recap\Golf Round Recap - Analysis.xlsx"

# ── Column indices (0-based, 29-column Rounds schema) ─────────────────────
C_DATE, C_COURSE, C_NOTE_TYPE, C_HOLE, C_PAR, C_DIST, C_SI = 0,1,2,3,4,5,6
C_SCORE, C_STROKES, C_PUTTS, C_PENALTIES                    = 7,8,9,10
C_FIR, C_GIR                                                = 11,12
C_TEE_CLUB, C_PICKUP, C_SENTIMENT                           = 13,14,15
C_DRIVER, C_WOODS, C_HYBRIDS, C_LONG_IRONS                  = 16,17,18,19
C_SHORT_IRONS, C_WEDGES, C_PUTTER                           = 20,21,22
C_HCP, C_TEE_COLOUR, C_RATING, C_SLOPE, C_WHS, C_NOTES      = 23,24,25,26,27,28

STRIKE_NUM   = {"Great": 4, "Good": 3, "Average": 2, "Poor": 1}
SENTIMENT_NUM = {"Positive": 5, "Neutral": 3, "Negative": 1}
SCORE_ORDER  = [
    "Hole in One","Albatross","Eagle","Birdie","Par",
    "Bogey","Double Bogey","Triple Bogey","Other","Pick Up",
]

# ── Styles ────────────────────────────────────────────────────────────────
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=10)
ALT_FILL = PatternFill("solid", fgColor="D9E8F5")
TOT_FILL = PatternFill("solid", fgColor="BDD7EE")
BOLD     = Font(bold=True, name="Calibri", size=11)
BODY     = Font(name="Calibri", size=11)

def hdr_row(ws, row, cols):
    for c, text in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=text)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[row].height = 28

def drow(ws, row, vals, bold=False, fill=None):
    for c, val in enumerate(vals, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.font = BOLD if bold else BODY
        if fill:
            cell.fill = fill

def set_widths(ws, widths):
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

def safe_avg(vals):
    clean = [v for v in vals if v is not None]
    return round(sum(clean) / len(clean), 1) if clean else None

def pct(n, d):
    return round(100 * n / d) if d else None


# ── Read source data ──────────────────────────────────────────────────────
wb_src = openpyxl.load_workbook(SRC, data_only=True)
ws_src = wb_src["Rounds"]

overall_rows, hole_rows = [], []
for row in ws_src.iter_rows(min_row=2, values_only=True):
    nt = row[C_NOTE_TYPE]
    if nt == "Overall":
        overall_rows.append(row)
    elif nt == "Hole":
        hole_rows.append(row)

overall_rows.sort(key=lambda r: r[C_DATE] or dt_date.min)
hole_rows.sort(key=lambda r: (r[C_DATE] or dt_date.min, r[C_HOLE] or 0))

wb = openpyxl.Workbook()


# ============================================================
# TAB 1: ROUND SUMMARY
# ============================================================
ws1 = wb.active
ws1.title = "Round Summary"
ws1.freeze_panes = "A2"

HDRS1 = [
    "Date", "Course", "Tee", "Score", "Par", "+/- Par", "Handicap Diff",
    "Playing HCP", "WHS Index", "Putts", "Penalties", "Sentiment",
    "Driver", "Woods", "Hybrids", "Long Irons", "Short Irons", "Wedges", "Putter",
    "Notes",
]
set_widths(ws1, [13,18,8,8,6,10,15, 13,11,8,11,12, 10,9,11,13,13,10,10, 55])
hdr_row(ws1, 1, HDRS1)

for i, row in enumerate(overall_rows):
    r  = i + 2
    fl = ALT_FILL if i % 2 == 0 else None
    score  = row[C_STROKES]
    par    = row[C_PAR]
    rating = row[C_RATING]
    slope  = row[C_SLOPE]
    vs_par   = (score - par) if (score and par) else None
    hcp_diff = round((score - rating) * 113 / slope, 1) if (score and rating and slope) else None

    drow(ws1, r, [
        row[C_DATE], row[C_COURSE], row[C_TEE_COLOUR],
        score, par, vs_par, hcp_diff,
        row[C_HCP], row[C_WHS], row[C_PUTTS], row[C_PENALTIES], row[C_SENTIMENT],
        row[C_DRIVER], row[C_WOODS], row[C_HYBRIDS],
        row[C_LONG_IRONS], row[C_SHORT_IRONS], row[C_WEDGES], row[C_PUTTER],
        row[C_NOTES],
    ], fill=fl)
    ws1.cell(row=r, column=1).number_format = "DD-MMM-YY"


# ============================================================
# TAB 2: HOLE STATS
# ============================================================
ws2 = wb.create_sheet("Hole Stats")
ws2.freeze_panes = "A2"

# Build (date, course) -> tee colour from Overall rows (hole rows don't store it)
tee_map = {(row[C_DATE], row[C_COURSE]): row[C_TEE_COLOUR] for row in overall_rows}

# Group hole rows by (Course, Tee Colour, Hole)
hole_data = defaultdict(list)
for row in hole_rows:
    tee = row[C_TEE_COLOUR] or tee_map.get((row[C_DATE], row[C_COURSE]), "?")
    key = (row[C_COURSE], tee, row[C_HOLE])
    hole_data[key].append(row)

# Snapshot of par/dist/si per hole (from first occurrence)
hole_meta = {
    key: {"par": rows[0][C_PAR], "dist": rows[0][C_DIST], "si": rows[0][C_SI]}
    for key, rows in hole_data.items()
}

HDRS2 = [
    "Course", "Tee", "Hole", "Par", "Dist (m)", "SI",
    "Rounds", "Avg Strokes", "Avg vs Par", "Best", "Worst",
    "FIR %", "GIR %", "Avg Putts", "Pen Hole %",
]
set_widths(ws2, [18,8,7,6,10,6, 9,13,12,8,9, 9,9,11,13])
hdr_row(ws2, 1, HDRS2)

sorted_keys = sorted(hole_data.keys(), key=lambda k: (k[0], k[1], k[2] or 0))

for i, key in enumerate(sorted_keys):
    course, tee, hole_num = key
    rows = hole_data[key]
    r    = i + 2
    fl   = ALT_FILL if i % 2 == 0 else None
    meta = hole_meta[key]
    par  = meta["par"]

    # Strokes -- exclude pick-up holes from average
    completed = [row for row in rows if row[C_PICKUP] != "Y" and row[C_STROKES] is not None]
    strokes_vals = [row[C_STROKES] for row in completed]
    avg_str  = safe_avg(strokes_vals)
    avg_vpar = round(avg_str - par, 1) if (avg_str and par) else None
    best     = min(strokes_vals) if strokes_vals else None
    worst    = max(strokes_vals) if strokes_vals else None

    # FIR -- par 4/5 only, Y/N recorded
    if par in (4, 5):
        fir_rec = [row for row in rows if row[C_FIR] in ("Y", "N")]
        fir_pct = pct(sum(1 for row in fir_rec if row[C_FIR] == "Y"), len(fir_rec))
    else:
        fir_pct = None

    # GIR -- all holes where Y/N recorded (blank on pick-ups)
    gir_rec = [row for row in rows if row[C_GIR] in ("Y", "N")]
    gir_pct = pct(sum(1 for row in gir_rec if row[C_GIR] == "Y"), len(gir_rec))

    # Putts -- where recorded
    putts_vals = [row[C_PUTTS] for row in rows if row[C_PUTTS] is not None]
    avg_putts  = safe_avg(putts_vals)

    # Penalties -- any hole with at least 1 penalty stroke
    pen_holes = sum(1 for row in rows if row[C_PENALTIES] and row[C_PENALTIES] > 0)
    pen_pct   = pct(pen_holes, len(rows))

    drow(ws2, r, [
        course, tee, hole_num, par, meta["dist"], meta["si"],
        len(rows),
        avg_str, avg_vpar, best, worst,
        fir_pct, gir_pct, avg_putts, pen_pct,
    ], fill=fl)

# Totals / averages row
r_tot = len(sorted_keys) + 2
all_completed = [row for row in hole_rows if row[C_PICKUP] != "Y" and row[C_STROKES] is not None]
all_strokes = [row[C_STROKES] for row in all_completed]
all_fir = [row for row in hole_rows if row[C_PAR] in (4,5) and row[C_FIR] in ("Y","N")]
all_gir = [row for row in hole_rows if row[C_GIR] in ("Y","N")]
all_putts = [row[C_PUTTS] for row in hole_rows if row[C_PUTTS] is not None]
all_pen   = [row for row in hole_rows if row[C_PENALTIES] and row[C_PENALTIES] > 0]

drow(ws2, r_tot, [
    "ALL", "", "TOTAL", "", "", "",
    len(overall_rows),                                        # rounds (not holes)
    safe_avg(all_strokes), None, None, None,
    pct(sum(1 for r in all_fir if r[C_FIR]=="Y"), len(all_fir)),
    pct(sum(1 for r in all_gir if r[C_GIR]=="Y"), len(all_gir)),
    safe_avg(all_putts),
    pct(len(all_pen), len(hole_rows)),
], bold=True, fill=TOT_FILL)


# ============================================================
# TAB 3: SCORING
# ============================================================
ws3 = wb.create_sheet("Scoring")
ws3.freeze_panes = "A2"

score_counts = defaultdict(int)
par_score    = defaultdict(lambda: defaultdict(int))
for row in hole_rows:
    label = row[C_SCORE]
    p     = row[C_PAR]
    if label:
        score_counts[label] += 1
    if label and p:
        par_score[p][label] += 1

HDRS3 = ["Score", "Total", "% of Holes", "Par 3", "Par 4", "Par 5"]
set_widths(ws3, [18, 10, 14, 10, 10, 10])
hdr_row(ws3, 1, HDRS3)

total_holes = len(hole_rows)
r = 2
for i, label in enumerate(SCORE_ORDER):
    count = score_counts.get(label, 0)
    if count == 0:
        continue
    fl = ALT_FILL if i % 2 == 0 else None
    drow(ws3, r, [
        label, count,
        round(100 * count / total_holes, 1) if total_holes else None,
        par_score[3].get(label) or None,
        par_score[4].get(label) or None,
        par_score[5].get(label) or None,
    ], fill=fl)
    r += 1

drow(ws3, r + 1, ["TOTAL", total_holes, 100, None, None, None], bold=True, fill=TOT_FILL)


# ============================================================
# TAB 4: SENTIMENT
# ============================================================
ws4 = wb.create_sheet("Sentiment")
ws4.freeze_panes = "A2"

HDRS4 = ["Course", "Tee", "Hole", "Par", "Rounds", "Positive", "Neutral", "Negative", "Most Common"]
set_widths(ws4, [18, 8, 7, 6, 9, 11, 10, 11, 15])
hdr_row(ws4, 1, HDRS4)

for i, key in enumerate(sorted_keys):
    course, tee, hole_num = key
    rows  = hole_data[key]
    r     = i + 2
    fl    = ALT_FILL if i % 2 == 0 else None
    sents = [row[C_SENTIMENT] for row in rows if row[C_SENTIMENT]]
    pos   = sents.count("Positive")
    neu   = sents.count("Neutral")
    neg   = sents.count("Negative")
    most  = max(["Positive","Neutral","Negative"], key=lambda s: sents.count(s)) if sents else None
    drow(ws4, r, [
        course, tee, hole_num, hole_meta[key]["par"],
        len(rows),
        pos or None, neu or None, neg or None, most,
    ], fill=fl)

# Round-level sentiment summary at bottom
r_sum = len(sorted_keys) + 3
ws4.cell(row=r_sum, column=1, value="ROUND SENTIMENT").font = HDR_FONT
ws4.cell(row=r_sum, column=1).fill = HDR_FILL
for col in range(2, 10):
    ws4.cell(row=r_sum, column=col).fill = HDR_FILL

r_sum += 1
hdr_row(ws4, r_sum, ["Date", "Course", "", "", "", "Positive", "Neutral", "Negative", "Overall"])

for i, row in enumerate(overall_rows):
    rs = r_sum + 1 + i
    fl = ALT_FILL if i % 2 == 0 else None
    drow(ws4, rs, [
        row[C_DATE], row[C_COURSE], "", "", "",
        None, None, None, row[C_SENTIMENT],
    ], fill=fl)
    ws4.cell(row=rs, column=1).number_format = "DD-MMM-YY"


# ── Save ──────────────────────────────────────────────────────────────────
wb.save(DEST)
print(f"Analysis saved: {DEST}")
print(f"  {len(overall_rows)} rounds | {len(hole_rows)} hole records")
print(f"  Tabs: Round Summary | Hole Stats | Scoring | Sentiment")
