"""
One-off data entry script -- Pupuke, 22 Feb 2026.
Clears all data rows and reloads this round with updated 29-column schema.
Delete this script once verified.
"""

import openpyxl
from openpyxl.styles import Font
from datetime import date

FILE = r"G:\My Drive\Project_Outputs\Golf Round Recap\Golf Round Recap.xlsx"
BODY_FONT = Font(name="Calibri", size=11)

wb = openpyxl.load_workbook(FILE)
ws = wb["Rounds"]

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for cell in row:
        cell.value = None

def wr(row_num, values):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.font = BODY_FONT

# ── Column order (29) ─────────────────────────────────────────────────────────
# 1  Date          8  Score         15 Pick Up       22 Wedges        28 WHS Index
# 2  Course        9  Strokes       16 Sentiment     23 Putter        29 Notes
# 3  Note Type    10  Putts         17 Driver        24 Handicap
# 4  Hole         11  Penalties     18 Woods         25 Tee Colour
# 5  Par          12  FIR           19 Hybrids       26 Course Rating
# 6  Distance     13  GIR           20 Long Irons    27 Slope
# 7  Stroke Index 14  Tee Club      21 Short Irons

D = date(2026, 2, 22)

# ── Overall row ───────────────────────────────────────────────────────────────
wr(2, [
    D, "Pupuke", "Overall", 0, 70, 5188, None,
    101, 101, None, None,
    None, None,          # FIR, GIR -- blank on Overall
    None, None,          # Tee Club, Pick Up
    "Negative",
    "Great", "Poor", "Average", "Poor", "Average", "Good", "Average",
    None, "White",
    68.2, 119, 24.4,     # Course Rating, Slope, WHS Index
    "Tee time 7:22am. Lovely late summer morning, a little cool. "
    "Next to no wind for most of the round, light winds on back nine. "
    "Dew on greens made them slower on the front nine, quickened a lot on the back nine. "
    "Great course conditions - firm ground, good bounce and roll out on shots. "
    "Disappointed with 101 - penalties hurt the score. "
    "Driver went well with good distance all day. Mid irons (6/7/8) were the main weakness - "
    "poor second shots after good drives is a wasted opportunity. "
    "Some good quality chips for gimmies. Putter has new grip, felt nice but work to do.",
])

# ── Hole data ─────────────────────────────────────────────────────────────────
HOLES = [
    # hole, par, dist, SI
    ( 1, 4, 300,  7),
    ( 2, 3, 139, 15),
    ( 3, 4, 335,  3),
    ( 4, 4, 304, 13),
    ( 5, 5, 431,  9),
    ( 6, 3, 165, 11),
    ( 7, 4, 363,  1),
    ( 8, 4, 333,  5),
    ( 9, 3, 147, 17),
    (10, 5, 422, 14),
    (11, 5, 398, 16),
    (12, 4, 362,  2),
    (13, 3, 167,  8),
    (14, 4, 285, 10),
    (15, 4, 299,  6),
    (16, 4, 238, 18),
    (17, 3, 143, 12),
    (18, 4, 357,  4),
]

STROKES = [5, 4, 7, 4, 6, 6, 7, 5, 4, 7, 6, 8, 6, 5, 5, 5, 5, 6]

def score_label(strokes, par, pickup):
    if pickup == "Y":
        return "Pick Up"
    return {-2:"Eagle",-1:"Birdie",0:"Par",1:"Bogey",
            2:"Double Bogey",3:"Triple Bogey"}.get(strokes - par, "Other")

# (tee_club, pickup, sentiment, putts, penalties, fir, gir, notes)
DETAIL = {
    1:  ("Driver",    "N", "Neutral",  None, None, "Y",  None,
         "Driver off tee. Sensible 8 iron second shot to lay up short of trouble. Bogey."),

    2:  ("8 Iron",    "N", "Negative", 3,    None, None, "Y",
         "8 iron off tee, landed on green. Three putt from the green - disappointing."),

    3:  ("Driver",    "N", "Negative", None, None, "Y",  "N",
         "Good driver off tee. Pulled 4 hybrid second shot - came up way below the green. "
         "Duffed first 60m chip. Bladed second 60m chip over the green. "
         "Duffed chip again. Putted out. Triple bogey - too many short game errors."),

    4:  ("Driver",    "N", "Positive", 2,    None, "Y",  "Y",
         "Big driver off tee. Good PW onto the green. Two putt. "
         "Best hole of the round - exactly the plan, solid par."),

    5:  ("Driver",    "N", "Neutral",  1,    None, "N",  "N",
         "Hooked driver off tee. 3 hybrid second - bit fadey but ok bounce. "
         "Tugged 8 iron left into netting - lucky bounce back into play. "
         "Good 80m wedge just right of pin. Great 54 chip for gimme putt. Bogey."),

    6:  ("5 Wood",    "N", "Negative", None, None, None, "N",
         "5 wood tee shot into bunker. Nightmare to get out - packed sand, kept hitting the lip. "
         "Basically a pick up score. Triple bogey."),

    7:  ("3 Hybrid",  "N", "Negative", None, None, "N",  "N",
         "3 hybrid off tee. Half-duffed 3 hybrid second - left just short of green. "
         "70m with 54 wedge. Bladed chip. Two or three putts. Triple bogey."),

    8:  ("Driver",    "N", "Neutral",  2,    None, "Y",  "N",
         "Monster driver - around 220m. Duffed 8 iron approach but lucky bounce kept in play. "
         "Good 54 wedge from 50m. Two putt. Bogey - decent recovery after the duff."),

    9:  ("6 Iron",    "N", "Neutral",  2,    None, None, "N",
         "6 iron off tee sprayed right. Slightly strong 54 wedge. Two putt. Bogey."),

    10: ("Driver",    "N", "Negative", 2,    1,    "N",  "N",
         "Strong driver hit a tree - ball out of play. "
         "Pulled 3 hybrid left - 1 penalty stroke to get out of trouble. "
         "50m 54 wedge - good distance but no spin, ball rolled back off the green. "
         "Two putt. Double bogey."),

    11: ("Driver",    "N", "Neutral",  3,    None, "Y",  "Y",
         "Monster drive. 3 hybrid second a bit fadey but safe right. "
         "50m 54 wedge - slightly hot, carried the front bunker clean. "
         "Three putt. Bogey on a par 5 - drive and approach good but let down by putting."),

    12: ("Driver",    "Y", "Negative", None, 1,    "N",  None,
         "Slicey driver - good distance but ended up on hole 13 fairway. "
         "Duffed 3 hybrid second. Terrible 7 iron - lost ball. Picked up. Dire hole."),

    13: ("5 Wood",    "Y", "Negative", None, None, None, None,
         "5 wood hooked badly left off tee. Hit provisional - found original ball but picked up."),

    14: ("Driver",    "N", "Neutral",  None, None, "Y",  "N",
         "Monster driver. GW from 85m - hit the green but rolled back off. "
         "Unlucky - good strike but no reward. Bogey."),

    15: ("Driver",    "N", "Positive", 2,    None, "Y",  "N",
         "Good driver - left side but good angle into the hole. "
         "7 iron faded right to greenside bunker. "
         "Nice bunker exit with flop shot. Two putt. Bogey."),

    16: ("3 Hybrid",  "N", "Negative", 2,    1,    "N",  "N",
         "Fadey 3 hybrid off tee - penalty. Good wedge. Two putt. Bogey."),

    17: ("8 Iron",    "N", "Negative", 3,    None, None, "N",
         "Duffed 8 iron tee shot on par 3. Good recovery wedge. "
         "Three putt - missed a short 4 footer. Double bogey."),

    18: ("Driver",    "N", "Negative", None, None, "N",  "N",
         "Big driver, bit fadey. Bit duffed 3 hybrid second. "
         "Bad wedge into bunker. Poor bunker shot. Putted out. Double bogey."),
}

for row_i, (hole, par, dist, si) in enumerate(HOLES, 3):
    strokes = STROKES[hole - 1]
    tee_club, pickup, sentiment, putts, penalties, fir, gir, notes = DETAIL[hole]
    score = score_label(strokes, par, pickup)
    wr(row_i, [
        D, "Pupuke", "Hole", hole, par, dist, si,
        score, strokes, putts, penalties,
        fir, gir,
        tee_club, pickup, sentiment,
        None, None, None, None, None, None, None,  # ball striking -- Overall only
        None, None,                                 # handicap, tee colour
        None, None, None,                           # course rating, slope, WHS -- Overall only
        notes,
    ])

wb.save(FILE)
print("Round loaded. 19 rows written (Overall + 18 holes).")
print("FIR/GIR, Course Rating 68.2, Slope 119, WHS Index 24.4 included.")
print("Delete this script once verified.")
