"""
Round data entry -- Purangi Golf & Country Club, 7 Mar 2026.
9-hole front nine. Appends course entry and round rows; safe to re-run (skips if already present).
"""

import openpyxl
from openpyxl.styles import Font
from datetime import date

FILE      = r"G:\My Drive\Project_Outputs\Golf Round Recap\Golf Round Recap.xlsx"
BODY_FONT = Font(name="Calibri", size=11)

wb    = openpyxl.load_workbook(FILE)
ws_r  = wb["Rounds"]
ws_c  = wb["Courses"]

# ── Column order (29) ──────────────────────────────────────────────────────────
# 1  Date          8  Score         15 Pick Up       22 Wedges        28 WHS Index
# 2  Course        9  Strokes       16 Sentiment     23 Putter        29 Notes
# 3  Note Type    10  Putts         17 Driver        24 Handicap
# 4  Hole         11  Penalties     18 Woods         25 Tee Colour
# 5  Par          12  FIR           19 Hybrids       26 Course Rating
# 6  Distance     13  GIR           20 Long Irons    27 Slope
# 7  Stroke Index 14  Tee Club      21 Short Irons

def wr(ws, values):
    row_num = ws.max_row + 1
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.font = BODY_FONT

# ── Guard: skip if round already present ──────────────────────────────────────
D = date(2026, 3, 7)
already_loaded = any(
    r[0].value == D and r[1].value == "Purangi" and r[2].value == "Overall"
    for r in ws_r.iter_rows(min_row=2)
    if r[0].value is not None
)
if already_loaded:
    print("Round already present — nothing written.")
    wb.close()
    exit()

# ── Guard: skip Courses entry if already present ───────────────────────────────
courses_loaded = any(
    r[0].value == "Purangi"
    for r in ws_c.iter_rows(min_row=2)
    if r[0].value is not None
)

# ── Courses tab — Purangi (White, front nine) ──────────────────────────────────
if not courses_loaded:
    for row in [
        ("Purangi", 1,       "White", 4, 274, 15, None, None, None),
        ("Purangi", 2,       "White", 4, 329,  5, None, None, None),
        ("Purangi", 3,       "White", 3, 150, 11, None, None, None),
        ("Purangi", 4,       "White", 5, 495,  1, None, None, None),
        ("Purangi", 5,       "White", 4, 357,  3, None, None, None),
        ("Purangi", 6,       "White", 4, 327,  7, None, None, None),
        ("Purangi", 7,       "White", 3, 115, 17, None, None, None),
        ("Purangi", 8,       "White", 5, 403,  9, None, None, None),
        ("Purangi", 9,       "White", 3, 159, 13, None, None, None),
        ("Purangi", "TOTAL", "White", 35, 2609, None, 67.0, 114,
         "9-hole front. Full 18: par 70, 5218m, NZCR 67, slope 114."),
    ]:
        wr(ws_c, row)
    print("Courses: Purangi added.")
else:
    print("Courses: Purangi already present — skipped.")

# ── Overall row ────────────────────────────────────────────────────────────────
wr(ws_r, [
    D, "Purangi", "Overall", 0, 35, 2609, None,
    43, 43, 18, 1,
    None, None,                        # FIR, GIR -- blank on Overall
    None, None,                        # Tee Club, Pick Up
    "Good",                            # Sentiment
    "Good", None, "Good",              # Driver, Woods, Hybrids
    "Average", None, "Average", "Average",  # Long Irons, Short Irons, Wedges, Putter
    24.6, "White",                     # Playing Handicap, Tee Colour
    67.0, 114, 24.6,                   # Course Rating, Slope, WHS Index
    "Short bag (D, 4H, 5i, 7i, 9i, 50\u00b0, 58\u00b0). Clear warm morning, "
    "dew on first 3 greens then dried. Firm track. Driver confident \u2014 "
    "good distance, slight fade developing. Fringe chipping good. Left putts short.",
])

# ── Hole data ──────────────────────────────────────────────────────────────────
# (hole, par, dist, SI, strokes, putts, pen, fir, gir, tee_club, notes)
HOLES = [
    (1, 4, 274, 15, 4, 2, 0, "Y",  "Y",  "Driver", None),
    (2, 4, 329,  5, 5, 2, 0, "Y",  "N",  "Driver", None),
    (3, 3, 150, 11, 4, 2, 0, None, "N",  "5i",     None),
    (4, 5, 495,  1, 5, 2, 0, "Y",  "Y",  "Driver", "Two cracking 4H approach shots (2nd & 3rd)"),
    (5, 4, 357,  3, 5, 2, 0, "N",  "N",  "Driver", None),
    (6, 4, 327,  7, 6, 2, 1, "N",  "N",  "Driver", "Great drive unlucky right, penalty"),
    (7, 3, 115, 17, 4, 2, 0, None, "N",  "9i",     "Good chip, close to saving par"),
    (8, 5, 403,  9, 6, 2, 0, "N",  "N",  "Driver", "Right into rough. Pulled 4H near creek. Thick pitch, second chip from 30 ok"),
    (9, 3, 159, 13, 4, 2, 0, None, "N",  "4H",     "4H short of green, good chip to 7ft, weak putt"),
]

for hole, par, dist, si, strokes, putts, pen, fir, gir, tee_club, notes in HOLES:
    wr(ws_r, [
        D, "Purangi", "Hole", hole, par, dist, si,
        strokes, strokes, putts, pen,
        fir, gir,
        tee_club, "N", None,
        None, None, None, None, None, None, None,  # ball striking -- Overall only
        None, "White",
        None, None, None,                           # Course Rating, Slope, WHS -- Overall only
        notes,
    ])

wb.save(FILE)
print(f"Round loaded. 10 rows written (Overall + 9 holes).")
